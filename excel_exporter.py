"""
Excel Export module for the Cultural Events Aggregator.

This module provides functionality to export filtered cultural events
to an Excel file with proper formatting and sorting.
"""

from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from parameters import OUTPUT_DIR, TIMEZONE


def _sort_events(events: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """
    Sort events by date and time.
    
    Args:
        events: List of event dictionaries.
        
    Returns:
        Sorted list of events (by date ascending, then time ascending).
    """
    def sort_key(event: Dict[str, str]) -> tuple:
        # Parse date (DD-MM-YYYY) to sortable format
        date_parts = event["date"].split("-")
        date_str = f"{date_parts[2]}{date_parts[1]}{date_parts[0]}"  # YYYYMMDD
        
        # Parse time (HH:MM) to sortable format
        time_str = event["time"].replace(":", "")  # HHMM
        
        return (date_str, time_str)
    
    return sorted(events, key=sort_key)


def _generate_filename() -> str:
    """
    Generate a filename with current timestamp in Warsaw timezone.
    
    Returns:
        Filename in format: events_YYYYMMDD_HHMMSS.xlsx
    """
    warsaw_tz = pytz.timezone(TIMEZONE)
    now = datetime.now(warsaw_tz)
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    return f"events_{timestamp}.xlsx"


def _auto_adjust_column_width(worksheet) -> None:
    """
    Automatically adjust column widths based on content.
    
    Args:
        worksheet: openpyxl worksheet object.
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                cell_length = len(str(cell.value)) if cell.value else 0
                if cell_length > max_length:
                    max_length = cell_length
            except (TypeError, AttributeError):
                pass
        
        # Add some padding and set max width
        adjusted_width = min(max_length + 2, 80)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def export_to_excel(
    events: List[Dict[str, str]],
    output_dir: Optional[str] = None,
) -> Optional[str]:
    """
    Export events to an Excel file.
    
    Args:
        events: List of event dictionaries with keys: date, time, event, link.
        output_dir: Directory to save the file (default: from parameters.py).
        
    Returns:
        Path to the created Excel file, or None if no events to export.
    """
    if not events:
        print("[WARNING] No events to export. Skipping Excel file creation.")
        return None
    
    # Use default output directory if not specified
    if output_dir is None:
        output_dir = OUTPUT_DIR
    
    # Create output directory if it doesn't exist
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    # Generate filename and full path
    filename = _generate_filename()
    file_path = output_path / filename
    
    # Sort events
    sorted_events = _sort_events(events)
    
    print(f"[INFO] Sorting {len(sorted_events)} events by date and time...")
    
    # Create workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Cultural Events"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    link_font = Font(color="0563C1", underline="single")
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    # Define headers
    headers = ["date", "time", "event", "link"]
    header_titles = ["Date", "Time", "Event", "Link"]
    
    # Write headers
    for col_idx, title in enumerate(header_titles, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Freeze header row
    worksheet.freeze_panes = "A2"
    
    # Write data rows
    for row_idx, event in enumerate(sorted_events, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = event.get(header, "")
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # Apply hyperlink style to link column
            if header == "link" and value:
                cell.font = link_font
                cell.hyperlink = value
    
    # Set specific column widths
    worksheet.column_dimensions["A"].width = 12  # Date
    worksheet.column_dimensions["B"].width = 8   # Time
    worksheet.column_dimensions["C"].width = 60  # Event
    worksheet.column_dimensions["D"].width = 50  # Link
    
    # Set row height for header
    worksheet.row_dimensions[1].height = 20
    
    # Save workbook
    workbook.save(file_path)
    
    print(f"[INFO] Generating Excel file...")
    
    return str(file_path)


if __name__ == "__main__":
    # Test the exporter
    sample_events = [
        {
            "date": "20-01-2026",
            "time": "10:00",
            "event": "Warsztaty plastyczne dla dzieci 4-7 lat",
            "link": "https://waw4free.pl/wydarzenie-12345",
        },
        {
            "date": "18-01-2026",
            "time": "15:00",
            "event": "Teatrzyk kukiełkowy dla najmłodszych",
            "link": "https://kultura.um.warszawa.pl/-/teatrzyk",
        },
        {
            "date": "18-01-2026",
            "time": "10:30",
            "event": "Czytanie bajek dla przedszkolaków",
            "link": "https://bigbookcafe.pl/event/bajki",
        },
    ]
    
    result = export_to_excel(sample_events)
    if result:
        print(f"\nTest file created: {result}")
